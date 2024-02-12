import openpyxl
import math

NUMBER_LIST = 3

COLUMN_DISCIPLINE = 0


wb = openpyxl.open("laba.xlsx")
table_2 = wb.worksheets[NUMBER_LIST]


print('''
      task_2
      ''')

def arithmeticAverageWeighted(indexDiscipline):
    numerator = 0
    denominator = 0
    for rating in range(1, 5):
        cellValue = table_2[indexDiscipline][rating].value
        if cellValue is None:
            cellValue = 0
        denominator += cellValue
        numerator += cellValue * (rating + 1)
    answer = numerator / denominator
    print(table_2[indexDiscipline][0].value, ":", round(answer, 2))
    return answer


for indexDiscipline in range(2, table_2.max_row + 1):
    arithmeticAverageWeighted(indexDiscipline)

print('''
      task_3
      ''')  

def weightedMeanSquare(indexDiscipline):
    numerator = 0
    denominator = 0
    for rating in range(1, 5):
        cellValue = table_2[indexDiscipline][rating].value
        if cellValue is None:
            cellValue = 0
        denominator += cellValue
        numerator += cellValue * pow(rating + 1, 2)
    answer = math.sqrt(numerator / denominator)
    print(table_2[indexDiscipline][0].value, ":", round(answer, 2))


for indexDiscipline in range(2, table_2.max_row + 1):
    weightedMeanSquare(indexDiscipline)

print('''
      task_4
      ''') 


def frequency(indexDiscipline):
    print(table_2[indexDiscipline][COLUMN_DISCIPLINE].value, ":")
    countRating = 0
    for rating in range(1, 5):
        cellValue = table_2[indexDiscipline][rating].value
        if cellValue is None:
            cellValue = 0
        countRating += cellValue
    for rating in range(1, 5):
        cellValue = table_2[indexDiscipline][rating].value
        if cellValue is None:
            cellValue = 0
        probability = cellValue / countRating
        print('"',rating + 1,'":', round(probability, 2))

for indexDiscipline in range(2, table_2.max_row + 1):
    frequency(indexDiscipline)